import csv
import io
import os
import cv2
import fitz
import traceback
import re
import json
import uuid
import tempfile
from pathlib import Path

import wordninja
from docling.document_converter import DocumentConverter, PdfFormatOption
from docling.datamodel.pipeline_options import PdfPipelineOptions
from transformers import AutoTokenizer

from vision_extraction import (
    extract_with_qwenvl,
    markdown_table_to_text,
    ensure_ollama_running,
)
from ultralytics import YOLO

# ---------------------------------------------------------------------------
# Model / converter setup
# ---------------------------------------------------------------------------

table_detector = YOLO("model/table_det.pt")

tokenizer = AutoTokenizer.from_pretrained("BAAI/bge-small-en-v1.5")


pipeline_options = PdfPipelineOptions()
pipeline_options.do_ocr = False
pipeline_options.do_table_structure = False
pipeline_options.generate_page_images = False
pipeline_options.images_scale = 1.0
converter = DocumentConverter(
    format_options={"pdf": PdfFormatOption(pipeline_options=pipeline_options)}
)

MAX_TOKENS = 450
TABLE_MAX_TOKENS = 350
TEXT_MAX_TOKENS = 350
TEXT_OVERLAP = 50

# Render scale for full-page scanned-OCR fallback. Lower than the 3x used
# for table/figure crops because a full page at 3x can exceed the vision
# model's context window (e.g. 4096 tokens). 1.5x keeps text legible while
# staying well under that limit for typical letter/A4 pages. Raise this if
# your Ollama model is configured with a larger num_ctx.
SCANNED_PAGE_SCALE = 1.5


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def token_count(text: str) -> int:
    return len(tokenizer.encode(text, add_special_tokens=False))


def clean_text(text) -> str:
    if text is None:
        return ""
    text = str(text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def fix_broken_spacing(text: str) -> str:
    text = clean_text(text)
    if not text:
        return ""
    words = text.split()
    if len(words) >= 3 or text.isupper():
        return text
    return " ".join(wordninja.split(text)).strip()


def get_item_bbox(item):
    """
    Get Docling item bbox in format:
    x1,y1,x2,y2
    """
    try:
        prov = item.prov[0]

        bbox = prov.bbox

        return (
            bbox.l,
            bbox.t,
            bbox.r,
            bbox.b,
        )

    except Exception:
        return None


def bbox_iou(box1, box2):
    """
    Calculate overlap ratio.
    box format:
    x1,y1,x2,y2
    """

    x1 = max(box1[0], box2[0])
    y1 = max(box1[1], box2[1])
    x2 = min(box1[2], box2[2])
    y2 = min(box1[3], box2[3])

    if x2 <= x1 or y2 <= y1:
        return 0

    intersection = (x2 - x1) * (y2 - y1)

    area1 = (box1[2] - box1[0]) * (box1[3] - box1[1])

    return intersection / area1


def is_heading_item(item) -> bool:
    return type(item).__name__ in {"TitleItem", "SectionHeaderItem", "HeadingItem"}


def is_picture_item(item) -> bool:
    return type(item).__name__ in {"PictureItem", "FigureItem"}


# ---------------------------------------------------------------------------
# PDF rendering / table detection / cropping
# ---------------------------------------------------------------------------


def render_page_to_image(pdf_path: str, page_no: int, scale: float = 3) -> str:
    """Render a PDF page (1-indexed) to a temp PNG at the given scale (default 3x)."""
    doc = fitz.open(pdf_path)
    try:
        page = doc[page_no - 1]
        pix = page.get_pixmap(matrix=fitz.Matrix(scale, scale), alpha=False)
        tmp = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
        tmp_path = tmp.name
        tmp.close()
        pix.save(tmp_path)
        return tmp_path
    finally:
        doc.close()


# def detect_tables(page_img: str) -> list[tuple[int, int, int, int]]:
#     results = table_detector.predict(page_img, conf=0.85, verbose=False, device="cpu")
#     boxes = []
#     for r in results:
#         for box in r.boxes.xyxy.cpu().numpy():
#             x1, y1, x2, y2 = map(int, box)
#             boxes.append((x1, y1, x2, y2))
#     return boxes


def detect_tables(
    page_img: str,
    scale: int = 3,
    conf: float = 0.85,
) -> list[dict]:

    results = table_detector.predict(page_img, conf=conf, verbose=False, device="cpu")

    tables = []

    for r in results:
        for box in r.boxes.xyxy.cpu().numpy():
            # YOLO bbox (image coordinate)
            ix1, iy1, ix2, iy2 = map(int, box)

            # Convert ke PDF coordinate
            px1 = int(ix1 / scale)
            py1 = int(iy1 / scale)
            px2 = int(ix2 / scale)
            py2 = int(iy2 / scale)

            tables.append(
                {
                    "image_bbox": (
                        ix1,
                        iy1,
                        ix2,
                        iy2,
                    ),
                    "pdf_bbox": (
                        px1,
                        py1,
                        px2,
                        py2,
                    ),
                }
            )

    return tables


def crop_table(img_path: str, bbox: tuple, pad: int = 5) -> str | None:
    img = cv2.imread(img_path)
    if img is None:
        return None

    h, w = img.shape[:2]
    x1, y1, x2, y2 = bbox

    # Guarantee correct ordering before padding/clamping.
    x1, x2 = sorted((x1, x2))
    y1, y2 = sorted((y1, y2))

    x1 = max(0, x1 - pad)
    y1 = max(0, y1 - pad)
    x2 = min(w, x2 + pad)
    y2 = min(h, y2 + pad)

    if x2 <= x1 or y2 <= y1:
        # Degenerate / out-of-bounds bbox -> nothing to crop.
        return None

    crop = img[y1:y2, x1:x2]
    if crop.size == 0:
        return None

    tmp = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
    tmp_path = tmp.name
    tmp.close()
    cv2.imwrite(tmp_path, crop)
    return tmp_path


# crop_table works on any bbox in image-pixel coordinates; reused for figures.
crop_region = crop_table


def pdf_bbox_to_image_bbox(
    bbox: tuple, page_height_pt: float, scale: int = 3
) -> tuple[int, int, int, int]:
    """
    Convert a Docling bbox (PDF coordinate space, origin bottom-left, y up)
    to an image-pixel bbox (origin top-left, y down) at the given render scale.

    Docling/PDF bbox fields commonly come as (l, t, r, b) where t > b in PDF
    space (t is closer to the top of the page = higher y value). We flip
    using page height, then sort so x1<x2 and y1<y2 for safe slicing.
    """
    x1, y1, x2, y2 = bbox

    # Flip Y: pdf_y -> image_y = (page_height - pdf_y)
    img_y1 = page_height_pt - y1
    img_y2 = page_height_pt - y2

    # Sort to guarantee top < bottom, left < right after flipping/scaling
    left, right = sorted((x1, x2))
    top, bottom = sorted((img_y1, img_y2))

    return (
        int(left * scale),
        int(top * scale),
        int(right * scale),
        int(bottom * scale),
    )


def get_pdf_page_height(pdf_path: str, page_no: int) -> float:
    """Return the height (in PDF points) of a 1-indexed page."""
    doc = fitz.open(pdf_path)
    try:
        return doc[page_no - 1].rect.height
    finally:
        doc.close()


def _remove_file(path: str | None) -> None:
    if path and os.path.exists(path):
        try:
            os.remove(path)
        except Exception:
            pass


# ---------------------------------------------------------------------------
# Chunking
# ---------------------------------------------------------------------------
def chunk_by_tokens(
    text_list: list[str],
    max_tokens: int = TEXT_MAX_TOKENS,
    overlap: int = TEXT_OVERLAP,
) -> list[str]:
    full_text = " ".join(clean_text(t) for t in text_list if clean_text(t))
    if not full_text:
        return []

    token_ids = tokenizer.encode(full_text, add_special_tokens=False)
    chunks = []
    start = 0
    while start < len(token_ids):
        end = min(start + max_tokens, len(token_ids))
        chunk_text = tokenizer.decode(token_ids[start:end], skip_special_tokens=True)
        chunks.append(chunk_text)
        if end == len(token_ids):
            break
        start += max_tokens - overlap
    return chunks


def chunk_table_text(table_text: str, max_tokens: int = TABLE_MAX_TOKENS) -> list[str]:
    """Split pipe-delimited table rows into token-safe chunks."""
    lines = table_text.splitlines()
    chunks: list[str] = []
    current: list[str] = []
    for line in lines:
        candidate = "\n".join(current + [line])
        if token_count(candidate) > max_tokens:
            if current:
                chunks.append("\n".join(current))
            current = [line]
        else:
            current.append(line)
    if current:
        chunks.append("\n".join(current))
    return chunks


# ---------------------------------------------------------------------------
# PDF batch splitting
# ---------------------------------------------------------------------------


def split_pdf_batches(pdf_path: str, batch_size: int = 3) -> list[str]:
    doc = fitz.open(pdf_path)
    temp_files = []
    for start in range(0, len(doc), batch_size):
        end = min(start + batch_size, len(doc))
        temp_doc = fitz.open()
        temp_doc.insert_pdf(doc, from_page=start, to_page=end - 1)
        tmp = tempfile.NamedTemporaryFile(suffix=".pdf", delete=False)
        tmp_path = tmp.name
        tmp.close()
        temp_doc.save(tmp_path)
        temp_doc.close()
        temp_files.append(tmp_path)
    doc.close()
    return temp_files


# ---------------------------------------------------------------------------
# Text flush
# ---------------------------------------------------------------------------


def flush_text_chunks(
    buffer_text: list[str],
    doc_id: str,
    source_pdf: str,
    page: int,
    section: str,
    all_chunks: list,
    global_index: int,
) -> int:
    if not buffer_text:
        return global_index

    chunks = chunk_by_tokens(buffer_text)
    for local_idx, chunk in enumerate(chunks):
        all_chunks.append(
            {
                "id": str(uuid.uuid4()),
                "doc_id": doc_id,
                "source_pdf": source_pdf,
                "page": page,
                "type": "text",
                "section": section,
                "content": chunk,
                "metadata": {
                    "chunk_index": global_index + local_idx,
                    "local_index": local_idx,
                    "token_count": token_count(chunk),
                },
            }
        )
    return global_index + len(chunks)


# ---------------------------------------------------------------------------
# Main pipeline
# ---------------------------------------------------------------------------


def process_pdf(pdf_path: str, batch_size: int = 3) -> dict:
    ensure_ollama_running()

    doc_id = str(uuid.uuid4())
    source_pdf = os.path.basename(pdf_path)

    batches = split_pdf_batches(pdf_path, batch_size)
    all_chunks: list[dict] = []
    global_chunk_index = 0

    for batch_idx, batch_file in enumerate(batches):
        batch_start_page = batch_idx * batch_size
        print(f"Batch {batch_idx + 1}/{len(batches)}")

        try:
            result = converter.convert(batch_file)
            doc = result.document

            buffer_text: list[str] = []
            current_page = 1
            section_stack: list[str] = []
            current_section = "Document"
            processed_table_pages: set[int] = set()

            page_table_boxes = {}
            # Tracks whether a page produced ANY content (text, table, figure)
            # so we can detect fully-scanned pages (no text layer at all).
            # Pre-initialized for every page in the batch — a page with zero
            # Docling items (pure scanned image) would otherwise never be
            # visited by the iterate_items() loop below.
            page_has_content: dict[int, bool] = {}
            try:
                _batch_doc = fitz.open(batch_file)
                for _p in range(1, _batch_doc.page_count + 1):
                    page_has_content[_p] = False
                _batch_doc.close()
            except Exception:
                traceback.print_exc()

            for item, level in doc.iterate_items():
                try:
                    # ── Track current page ──────────────────────────────────
                    try:
                        current_page = item.prov[0].page_no
                    except Exception:
                        pass

                    print(f"[DEBUG] item={type(item).__name__} page={current_page} level={level} text={getattr(item, 'text', None)!r:.80}")

                    page_has_content.setdefault(current_page, False)

                    # ── Table detection (once per page) ─────────────────────
                    if current_page not in processed_table_pages:
                        page_img = None

                        try:
                            absolute_page = batch_start_page + current_page

                            page_img = render_page_to_image(pdf_path, absolute_page)
                            detected_tables = detect_tables(page_img)

                            page_table_boxes[current_page] = [
                                table["pdf_bbox"] for table in detected_tables
                            ]

                            # ==============================
                            # Extract table content
                            # ==============================

                            for table in detected_tables:
                                image_bbox = table["image_bbox"]

                                crop = crop_table(page_img, image_bbox)

                                if crop is None:
                                    print(
                                        f"[Crop FAILED] table "
                                        f"page={current_page} bbox={image_bbox}"
                                    )
                                    continue

                                try:
                                    markdown = extract_with_qwenvl(crop, task="table")

                                finally:
                                    _remove_file(crop)

                                if not markdown:
                                    print(
                                        f"[QwenVL FAILED] "
                                        f"page={current_page} "
                                        f"bbox={image_bbox}"
                                    )
                                    continue

                                table_text = markdown_table_to_text(markdown)

                                if not table_text.strip():
                                    continue

                                table_chunks = chunk_table_text(table_text)

                                for chunk in table_chunks:
                                    all_chunks.append(
                                        {
                                            "id": str(uuid.uuid4()),
                                            "doc_id": doc_id,
                                            "source_pdf": source_pdf,
                                            "page": current_page,
                                            "type": "table",
                                            "section": current_section,
                                            "content": chunk,
                                            "metadata": {
                                                "chunk_index": global_chunk_index,
                                                "token_count": token_count(chunk),
                                                # simpan image bbox
                                                # karena ini lokasi crop
                                                "bbox": list(image_bbox),
                                            },
                                        }
                                    )

                                    global_chunk_index += 1
                                    page_has_content[current_page] = True

                        finally:
                            _remove_file(page_img)

                        processed_table_pages.add(current_page)

                    # ── Section heading ─────────────────────────────────────
                    if is_heading_item(item):
                        heading = clean_text(getattr(item, "text", ""))
                        if heading:
                            if buffer_text:
                                global_chunk_index = flush_text_chunks(
                                    buffer_text,
                                    doc_id,
                                    source_pdf,
                                    current_page,
                                    current_section,
                                    all_chunks,
                                    global_chunk_index,
                                )
                                buffer_text = []

                            while len(section_stack) > level:
                                section_stack.pop()
                            section_stack.append(heading)
                            current_section = " > ".join(section_stack)
                            page_has_content[current_page] = True
                        continue

                    # ── Figure / chart / diagram (non-table picture) ────────
                    if is_picture_item(item):
                        item_bbox = get_item_bbox(item)

                        if item_bbox is None:
                            continue

                        absolute_page = batch_start_page + current_page

                        try:
                            page_height_pt = get_pdf_page_height(pdf_path, absolute_page)
                        except Exception:
                            traceback.print_exc()
                            continue

                        image_bbox = pdf_bbox_to_image_bbox(item_bbox, page_height_pt)

                        # Skip figures that overlap an already-detected table
                        # (handled by the table pipeline above). Both bboxes
                        # are now in the same image-pixel coordinate space.
                        overlaps_table = False
                        for table_bbox in page_table_boxes.get(current_page, []):
                            if bbox_iou(image_bbox, table_bbox) > 0.3:
                                overlaps_table = True
                                break

                        if overlaps_table:
                            continue

                        fig_page_img = None
                        crop = None

                        try:
                            fig_page_img = render_page_to_image(pdf_path, absolute_page)
                            crop = crop_region(fig_page_img, image_bbox)

                            if crop is None:
                                print(
                                    f"[Crop FAILED] figure "
                                    f"page={current_page} bbox={item_bbox}"
                                )
                                continue

                            description = extract_with_qwenvl(crop, task="explain")

                            if not description or not description.strip():
                                print(
                                    f"[QwenVL FAILED] figure "
                                    f"page={current_page} bbox={item_bbox}"
                                )
                                continue

                            all_chunks.append(
                                {
                                    "id": str(uuid.uuid4()),
                                    "doc_id": doc_id,
                                    "source_pdf": source_pdf,
                                    "page": current_page,
                                    "type": "figure",
                                    "section": current_section,
                                    "content": description.strip(),
                                    "metadata": {
                                        "chunk_index": global_chunk_index,
                                        "token_count": token_count(description),
                                        "bbox": list(image_bbox),
                                    },
                                }
                            )

                            global_chunk_index += 1
                            page_has_content[current_page] = True

                        finally:
                            _remove_file(crop)
                            _remove_file(fig_page_img)

                        continue

                    # ── Regular text ────────────────────────────────────────
                    if hasattr(item, "text"):
                        text = clean_text(item.text)

                        if not text:
                            continue

                        item_bbox = get_item_bbox(item)

                        if item_bbox:
                            skip = False

                            for table_bbox in page_table_boxes.get(current_page, []):
                                if bbox_iou(item_bbox, table_bbox) > 0.3:
                                    skip = True
                                    break

                            if skip:
                                continue

                        buffer_text.append(text)
                        page_has_content[current_page] = True

                except Exception:
                    traceback.print_exc()

            # Flush any remaining text buffer at end of batch
            if buffer_text:
                global_chunk_index = flush_text_chunks(
                    buffer_text,
                    doc_id,
                    source_pdf,
                    current_page,
                    current_section,
                    all_chunks,
                    global_chunk_index,
                )

            # ── Scanned-page fallback ────────────────────────────────────────
            # Any page in this batch that produced zero text/table/figure
            # content is likely a scanned image with no text layer at all.
            # Render it and run full-page OCR via QwenVL.
            #
            # NOTE: full-page renders use a lower scale than table/figure
            # crops. The model's context window (e.g. 4096 tokens) can be
            # exceeded by a full page rendered at 3x — a crop of one table
            # or figure is much smaller than a full page, so it doesn't hit
            # this limit. Tune SCANNED_PAGE_SCALE if your model has a larger
            # context window and you want higher-fidelity OCR.
            for local_page, has_content in page_has_content.items():
                if has_content:
                    continue

                absolute_page = batch_start_page + local_page
                scan_img = None

                try:
                    scan_img = render_page_to_image(
                        pdf_path, absolute_page, scale=SCANNED_PAGE_SCALE
                    )
                    ocr_text = extract_with_qwenvl(scan_img, task="document")

                    if not ocr_text or not ocr_text.strip():
                        print(f"[QwenVL FAILED] scanned page={local_page}")
                        continue

                    ocr_chunks = chunk_by_tokens([ocr_text])

                    for local_idx, chunk in enumerate(ocr_chunks):
                        all_chunks.append(
                            {
                                "id": str(uuid.uuid4()),
                                "doc_id": doc_id,
                                "source_pdf": source_pdf,
                                "page": local_page,
                                "type": "scanned_ocr",
                                "section": "Document",
                                "content": chunk,
                                "metadata": {
                                    "chunk_index": global_chunk_index + local_idx,
                                    "local_index": local_idx,
                                    "token_count": token_count(chunk),
                                },
                            }
                        )

                    global_chunk_index += len(ocr_chunks)

                finally:
                    _remove_file(scan_img)

        except Exception:
            traceback.print_exc()

        finally:
            _remove_file(batch_file)

    print(f"Total chunks: {len(all_chunks)}")
    return {"doc_id": doc_id, "source_pdf": source_pdf, "chunks": all_chunks}


# ---------------------------------------------------------------------------
# Supported formats
# ---------------------------------------------------------------------------

# Formats routed through Docling (text + heading pipeline, no YOLO table detection)
_DOCLING_EXTS = {".docx", ".pptx", ".html", ".htm", ".md", ".txt", ".asciidoc", ".adoc"}

# Formats handled as flat tabular data
_TABULAR_EXTS = {".xlsx", ".xlsm", ".csv"}


# ---------------------------------------------------------------------------
# Non-PDF extraction helpers
# ---------------------------------------------------------------------------


def _iter_docling_doc(doc, source_name: str, doc_id: str) -> dict:
    """
    Run the heading + text pipeline on an already-converted Docling document.
    No YOLO / QwenVL — only text items. Returns the standard result dict.
    Page numbers are kept where Docling provides them (1-indexed); non-paged
    formats (DOCX, PPTX) get page 1 as fallback.
    """
    all_chunks: list[dict] = []
    global_chunk_index = 0

    buffer_text: list[str] = []
    current_page = 1
    section_stack: list[str] = []
    current_section = "Document"

    for item, level in doc.iterate_items():
        try:
            try:
                current_page = item.prov[0].page_no
            except Exception:
                pass

            if is_heading_item(item):
                heading = clean_text(getattr(item, "text", ""))
                if heading:
                    if buffer_text:
                        global_chunk_index = flush_text_chunks(
                            buffer_text,
                            doc_id,
                            source_name,
                            current_page,
                            current_section,
                            all_chunks,
                            global_chunk_index,
                        )
                        buffer_text = []

                    while len(section_stack) > level:
                        section_stack.pop()
                    section_stack.append(heading)
                    current_section = " > ".join(section_stack)
                continue

            if hasattr(item, "text"):
                text = clean_text(item.text)
                if text:
                    buffer_text.append(text)

        except Exception:
            traceback.print_exc()

    if buffer_text:
        global_chunk_index = flush_text_chunks(
            buffer_text,
            doc_id,
            source_name,
            current_page,
            current_section,
            all_chunks,
            global_chunk_index,
        )

    return {"doc_id": doc_id, "source_pdf": source_name, "chunks": all_chunks}


def _process_docling_format(file_path: str) -> dict:
    """Convert DOCX / PPTX / HTML / MD / TXT via Docling then run text pipeline."""
    doc_id = str(uuid.uuid4())
    source_name = os.path.basename(file_path)
    print(f"Processing via Docling text pipeline: {source_name}")

    result = converter.convert(file_path)
    return _iter_docling_doc(result.document, source_name, doc_id)


def _process_xlsx(file_path: str) -> dict:
    """Read every sheet in an Excel file as tabular data and chunk by token limit."""
    from openpyxl import load_workbook

    doc_id = str(uuid.uuid4())
    source_name = os.path.basename(file_path)
    all_chunks: list[dict] = []
    global_chunk_index = 0

    print(f"Processing Excel: {source_name}")

    wb = load_workbook(file_path, read_only=True, data_only=True)
    for sheet in wb.worksheets:
        rows: list[str] = []
        for row in sheet.iter_rows(values_only=True):
            cells = [str(c).strip() if c is not None else "" for c in row]
            if any(cells):
                rows.append(" | ".join(cells))

        if not rows:
            continue

        table_text = "\n".join(rows)
        section = sheet.title or "Sheet"

        for chunk in chunk_table_text(table_text):
            all_chunks.append(
                {
                    "id": str(uuid.uuid4()),
                    "doc_id": doc_id,
                    "source_pdf": source_name,
                    "page": 1,
                    "type": "table",
                    "section": section,
                    "content": chunk,
                    "metadata": {
                        "chunk_index": global_chunk_index,
                        "token_count": token_count(chunk),
                        "sheet": section,
                    },
                }
            )
            global_chunk_index += 1

    wb.close()
    print(f"Total chunks: {len(all_chunks)}")
    return {"doc_id": doc_id, "source_pdf": source_name, "chunks": all_chunks}


def _process_csv(file_path: str) -> dict:
    """Read a CSV file as tabular data and chunk by token limit."""
    doc_id = str(uuid.uuid4())
    source_name = os.path.basename(file_path)
    all_chunks: list[dict] = []
    global_chunk_index = 0

    print(f"Processing CSV: {source_name}")

    with open(file_path, encoding="utf-8", errors="replace", newline="") as f:
        reader = csv.reader(f)
        rows = [" | ".join(cell.strip() for cell in row) for row in reader if any(row)]

    if not rows:
        return {"doc_id": doc_id, "source_pdf": source_name, "chunks": []}

    table_text = "\n".join(rows)
    for chunk in chunk_table_text(table_text):
        all_chunks.append(
            {
                "id": str(uuid.uuid4()),
                "doc_id": doc_id,
                "source_pdf": source_name,
                "page": 1,
                "type": "table",
                "section": "Document",
                "content": chunk,
                "metadata": {
                    "chunk_index": global_chunk_index,
                    "token_count": token_count(chunk),
                },
            }
        )
        global_chunk_index += 1

    print(f"Total chunks: {len(all_chunks)}")
    return {"doc_id": doc_id, "source_pdf": source_name, "chunks": all_chunks}


# ---------------------------------------------------------------------------
# Unified entry point
# ---------------------------------------------------------------------------


def process_file(file_path: str, batch_size: int = 3) -> dict:
    """
    Universal entry point. Routes to the correct extraction pipeline based on
    file extension:

      .pdf                    -> full multimodal PDF pipeline:
                                   - Docling text + headings
                                   - YOLO + QwenVL table extraction
                                   - Docling PictureItem + QwenVL figure
                                     description (charts, diagrams, photos)
                                   - PyMuPDF render + QwenVL full-page OCR
                                     fallback for scanned pages with no
                                     text layer at all
      .docx .pptx .html .md
      .txt .asciidoc .adoc   -> Docling text + heading pipeline
      .xlsx .xlsm            -> Excel sheet reader -> tabular chunks
      .csv                   -> CSV reader -> tabular chunks

    Returns the standard result dict: {doc_id, source_pdf, chunks}.
    Chunk "type" values: "text", "table", "figure", "scanned_ocr".
    Raises ValueError for unsupported extensions.
    """
    ext = Path(file_path).suffix.lower()

    if ext == ".pdf":
        return process_pdf(file_path, batch_size=batch_size)

    if ext in _DOCLING_EXTS:
        return _process_docling_format(file_path)

    if ext in {".xlsx", ".xlsm"}:
        return _process_xlsx(file_path)

    if ext == ".csv":
        return _process_csv(file_path)

    raise ValueError(
        f"Unsupported file format: '{ext}'. "
        f"Supported: .pdf, {', '.join(sorted(_DOCLING_EXTS | _TABULAR_EXTS))}"
    )


# ---------------------------------------------------------------------------
# Output helpers
# ---------------------------------------------------------------------------


def save_jsonl(result: dict, output_path: str) -> None:
    with open(output_path, "w", encoding="utf-8") as f:
        for chunk in result["chunks"]:
            f.write(json.dumps(chunk, ensure_ascii=False) + "\n")
    print(f"Saved: {output_path}")


def preview_chunks(result: dict, n: int = 5) -> None:
    print("\n")
    print("=" * 80)
    print("EXTRACTION PREVIEW")
    print("=" * 80)
    for i, chunk in enumerate(result["chunks"][:n]):
        print(f"\nChunk {i + 1}")
        print(f"Page   : {chunk['page']}")
        print(f"Type   : {chunk['type']}")
        print(f"Section: {chunk['section']}")
        print(f"Tokens : {chunk['metadata']['token_count']}")
        print(chunk["content"][:300])
        print("-" * 80)


if __name__ == "__main__":
    FILE_PATH = r"C:\Users\asus\OneDrive\Documents\docs\img_sample.pdf"
    OUTPUT = "img_sample.jsonl"
    result = process_file(FILE_PATH, batch_size=2)
    preview_chunks(result, n=10)
    save_jsonl(result, OUTPUT)
    print("\nDONE")